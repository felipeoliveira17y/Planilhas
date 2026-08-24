from pathlib import Path
from openpyxl import load_workbook


ARQUIVO_EXCEL = Path(__file__).parent / "SistemaAgronomia.xlsx"


def abrir_excel():
    if not ARQUIVO_EXCEL.exists():
        raise FileNotFoundError(
            f"Arquivo Excel não encontrado:\n{ARQUIVO_EXCEL}"
        )

    return load_workbook(ARQUIVO_EXCEL)


def encontrar_tabela(wb, nome_tabela):

    for ws in wb.worksheets:

        if nome_tabela in ws.tables:

            tabela = ws.tables[nome_tabela]

            return ws, tabela

    raise ValueError(
        f"A tabela '{nome_tabela}' não foi encontrada."
    )


def ler_tabela(nome_tabela):

    wb = abrir_excel()

    ws, tabela = encontrar_tabela(
        wb,
        nome_tabela
    )

    dados = list(ws[tabela.ref])

    cabecalho = [
        celula.value
        for celula in dados[0]
    ]

    resultado = []

    for linha in dados[1:]:

        valores = [
            celula.value
            for celula in linha
        ]

        if not any(
            valor is not None
            for valor in valores
        ):
            continue

        resultado.append(
            dict(zip(cabecalho, valores))
        )

    return resultado


def gerar_id_cliente():

    clientes = ler_tabela("tbClientes")

    maior_id = 0

    for cliente in clientes:

        id_cliente = cliente.get("ID_CLIENTE")

        if not id_cliente:
            continue

        try:
            numero = int(
                id_cliente.split("-")[1]
            )

            maior_id = max(
                maior_id,
                numero
            )

        except (ValueError, IndexError):
            continue

    proximo = maior_id + 1

    return f"CLI-{proximo:03d}"


def adicionar_cliente(dados):

    wb = abrir_excel()

    ws, tabela = encontrar_tabela(
        wb,
        "tbClientes"
    )

    # Descobre os limites atuais da tabela
    inicio, fim = tabela.ref.split(":")

    # Coluna e linha inicial
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row

    # Coluna e linha final
    coluna_final = ws[fim].column
    linha_final = ws[fim].row

    # A próxima linha será logo abaixo da tabela
    proxima_linha = linha_final + 1

    # Obtém os cabeçalhos da tabela
    cabecalho = []

    for coluna in range(
        coluna_inicial,
        coluna_final + 1
    ):
        cabecalho.append(
            ws.cell(
                linha_inicial,
                coluna
            ).value
        )

    # Insere os dados
    for coluna, nome_coluna in enumerate(
        cabecalho,
        start=coluna_inicial
    ):

        valor = dados.get(
            nome_coluna,
            ""
        )

        ws.cell(
            proxima_linha,
            coluna,
            valor
        )

    # Atualiza o intervalo da tabela
    nova_referencia = (
        f"{inicio}:"
        f"{ws.cell(proxima_linha, coluna_final).coordinate}"
    )

    tabela.ref = nova_referencia

    # Salva o Excel
    wb.save(ARQUIVO_EXCEL)

    # Expande a tabela
    inicio = tabela.ref.split(":")[0]
    fim_coluna = ws.cell(
        1,
        coluna_inicial + len(cabecalho) - 1
    ).column_letter

    tabela.ref = (
        f"{inicio}:{fim_coluna}{proxima_linha}"
    )

    wb.save(ARQUIVO_EXCEL)

def gerar_id_propriedade():
    propriedades = ler_tabela("tbPropriedades")
    maior_id = 0

    for prop in propriedades:
        id_prop = prop.get("ID_PROPRIEDADE")
        if not id_prop:
            continue
        try:
            numero = int(id_prop.split("-")[1])
            maior_id = max(maior_id, numero)
        except (ValueError, IndexError):
            continue

    proximo = maior_id + 1
    return f"PROP-{proximo:03d}"


def adicionar_propriedade(dados):
    wb = abrir_excel()
    ws, tabela = encontrar_tabela(wb, "tbPropriedades")

    inicio, fim = tabela.ref.split(":")
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row
    coluna_final = ws[fim].column
    linha_final = ws[fim].row

    proxima_linha = linha_final + 1

    cabecalho = []
    for coluna in range(coluna_inicial, coluna_final + 1):
        cabecalho.append(ws.cell(linha_inicial, coluna).value)

    for coluna, nome_coluna in enumerate(cabecalho, start=coluna_inicial):
        valor = dados.get(nome_coluna, "")
        ws.cell(proxima_linha, coluna, valor)

    nova_referencia = f"{inicio}:{ws.cell(proxima_linha, coluna_final).coordinate}"
    tabela.ref = nova_referencia

    wb.save(ARQUIVO_EXCEL)

def gerar_id_servico():
    servicos = ler_tabela("tbServicos")
    maior_id = 0

    for serv in servicos:
        id_serv = serv.get("ID_SERVICO")
        if not id_serv:
            continue
        try:
            numero = int(id_serv.split("-")[1])
            maior_id = max(maior_id, numero)
        except (ValueError, IndexError):
            continue

    proximo = maior_id + 1
    return f"SERV-{proximo:03d}"


def adicionar_servico(dados):
    wb = abrir_excel()
    ws, tabela = encontrar_tabela(wb, "tbServicos")

    inicio, fim = tabela.ref.split(":")
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row
    coluna_final = ws[fim].column
    linha_final = ws[fim].row

    proxima_linha = linha_final + 1

    cabecalho = []
    for coluna in range(coluna_inicial, coluna_final + 1):
        cabecalho.append(ws.cell(linha_inicial, coluna).value)

    for coluna, nome_coluna in enumerate(cabecalho, start=coluna_inicial):
        valor = dados.get(nome_coluna, "")
        ws.cell(proxima_linha, coluna, valor)

    nova_referencia = f"{inicio}:{ws.cell(proxima_linha, coluna_final).coordinate}"
    tabela.ref = nova_referencia

    wb.save(ARQUIVO_EXCEL)

def gerar_id_equipamento():
    equipamentos = ler_tabela("tbEquipamentos")
    maior_id = 0
    for eq in equipamentos:
        id_eq = eq.get("ID_EQUIPAMENTO")
        if not id_eq:
            continue
        try:
            numero = int(id_eq.split("-")[1])
            maior_id = max(maior_id, numero)
        except (ValueError, IndexError):
            continue
    return f"EQP-{maior_id + 1:03d}"


def adicionar_equipamento(dados):
    wb = abrir_excel()
    ws, tabela = encontrar_tabela(wb, "tbEquipamentos")
    inicio, fim = tabela.ref.split(":")
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row
    coluna_final = ws[fim].column
    linha_final = ws[fim].row
    proxima_linha = linha_final + 1

    cabecalho = [ws.cell(linha_inicial, c).value for c in range(coluna_inicial, coluna_final + 1)]
    for coluna, nome_coluna in enumerate(cabecalho, start=coluna_inicial):
        ws.cell(proxima_linha, coluna, dados.get(nome_coluna, ""))

    tabela.ref = f"{inicio}:{ws.cell(proxima_linha, coluna_final).coordinate}"
    wb.save(ARQUIVO_EXCEL)


def gerar_id_receita():
    receitas = ler_tabela("tbReceitas")
    maior_id = 0
    for rec in receitas:
        id_rec = rec.get("ID_RECEITA")
        if not id_rec:
            continue
        try:
            numero = int(id_rec.split("-")[1])
            maior_id = max(maior_id, numero)
        except (ValueError, IndexError):
            continue
    return f"REC-{maior_id + 1:03d}"


def adicionar_receita(dados):
    wb = abrir_excel()
    ws, tabela = encontrar_tabela(wb, "tbReceitas")
    inicio, fim = tabela.ref.split(":")
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row
    coluna_final = ws[fim].column
    linha_final = ws[fim].row
    proxima_linha = linha_final + 1

    cabecalho = [ws.cell(linha_inicial, c).value for c in range(coluna_inicial, coluna_final + 1)]
    for coluna, nome_coluna in enumerate(cabecalho, start=coluna_inicial):
        ws.cell(proxima_linha, coluna, dados.get(nome_coluna, ""))

    tabela.ref = f"{inicio}:{ws.cell(proxima_linha, coluna_final).coordinate}"
    wb.save(ARQUIVO_EXCEL)


def gerar_id_despesa():
    despesas = ler_tabela("tbDespesas")
    maior_id = 0
    for desp in despesas:
        id_desp = desp.get("ID_DESPESA")
        if not id_desp:
            continue
        try:
            numero = int(id_desp.split("-")[1])
            maior_id = max(maior_id, numero)
        except (ValueError, IndexError):
            continue
    return f"DESP-{maior_id + 1:03d}"


def adicionar_despesa(dados):
    wb = abrir_excel()
    ws, tabela = encontrar_tabela(wb, "tbDespesas")
    inicio, fim = tabela.ref.split(":")
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row
    coluna_final = ws[fim].column
    linha_final = ws[fim].row
    proxima_linha = linha_final + 1

    cabecalho = [ws.cell(linha_inicial, c).value for c in range(coluna_inicial, coluna_final + 1)]
    for coluna, nome_coluna in enumerate(cabecalho, start=coluna_inicial):
        ws.cell(proxima_linha, coluna, dados.get(nome_coluna, ""))

    tabela.ref = f"{inicio}:{ws.cell(proxima_linha, coluna_final).coordinate}"
    wb.save(ARQUIVO_EXCEL)

def gerar_id_documento():
    documentos = ler_tabela("tbDocumentos")
    maior_id = 0
    for doc in documentos:
        id_doc = doc.get("ID_DOCUMENTO")
        if not id_doc:
            continue
        try:
            numero = int(id_doc.split("-")[1])
            maior_id = max(maior_id, numero)
        except (ValueError, IndexError):
            continue
    return f"DOC-{maior_id + 1:03d}"


def adicionar_documento(dados):
    wb = abrir_excel()
    ws, tabela = encontrar_tabela(wb, "tbDocumentos")
    inicio, fim = tabela.ref.split(":")
    coluna_inicial = ws[inicio].column
    linha_inicial = ws[inicio].row
    coluna_final = ws[fim].column
    linha_final = ws[fim].row
    proxima_linha = linha_final + 1

    cabecalho = [ws.cell(linha_inicial, c).value for c in range(coluna_inicial, coluna_final + 1)]
    for coluna, nome_coluna in enumerate(cabecalho, start=coluna_inicial):
        ws.cell(proxima_linha, coluna, dados.get(nome_coluna, ""))

    tabela.ref = f"{inicio}:{ws.cell(proxima_linha, coluna_final).coordinate}"
    wb.save(ARQUIVO_EXCEL)